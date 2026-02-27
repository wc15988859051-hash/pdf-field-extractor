#!/usr/bin/env python3
"""
Excel 导出脚本
基于模板格式填充数据，每组数据占 6 行
支持增量更新：如果 Excel 已存在，自动去重并追加新数据
自动合并单元格，保持格式一致
"""

import sys
import json
import os
import shutil
from typing import List, Dict, Any, Tuple

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请执行: pip install openpyxl")
    sys.exit(1)


# 字段映射关系：原字段名 → 新字段名
FIELD_MAPPING = {
    "Article": "color code",
    "Order Reference": "PO",
    "Colour Name": "color",
    "GBP Retail Price": "sell",
    "Collection": "style no",
    "Design Number": "style code",
    "Ex Port Date": "Ex-date",
    "Total": "quantity",
    "Unit Price": "unit price",
    "Line Value": "total",
    "Product Name": "style name"
}


def map_fields(data: List[Dict[str, Any]], pdf_filename: str = None) -> List[Dict[str, Any]]:
    """
    根据映射关系转换字段名，并添加文件名列

    Args:
        data: 原始数据列表，每个元素是一个字典
        pdf_filename: PDF 文件名（可选）

    Returns:
        转换后的数据列表
    """
    mapped_data = []
    for item in data:
        mapped_item = {}
        for old_key, value in item.items():
            # 查找映射关系
            new_key = FIELD_MAPPING.get(old_key, old_key)
            mapped_item[new_key] = value

        # 添加文件名列（如果提供了文件名）
        if pdf_filename:
            mapped_item["原PDF名称"] = pdf_filename

        mapped_data.append(mapped_item)
    return mapped_data


def merge_cells_for_group(ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int) -> None:
    """
    为一组数据合并单元格

    Args:
        ws: 工作表对象
        start_row: 起始行号
    """
    # 序号（列 1）：合并 5 行
    ws.merge_cells(f"A{start_row}:A{start_row+4}")

    # style no（列 3）：合并 5 行
    ws.merge_cells(f"C{start_row}:C{start_row+4}")

    # style name（列 4）：合并 5 行
    ws.merge_cells(f"D{start_row}:D{start_row+4}")

    # color（列 5）：合并 5 行
    ws.merge_cells(f"E{start_row}:E{start_row+4}")

    # unit price（列 6）：前 3 行合并（行 1-3）
    ws.merge_cells(f"F{start_row}:F{start_row+2}")
    # unit price（列 6）：后 2 行合并（行 4-5）
    ws.merge_cells(f"F{start_row+3}:F{start_row+4}")

    # quantity（列 7）：合并 5 行
    ws.merge_cells(f"G{start_row}:G{start_row+4}")

    # amount（列 8）：合并 5 行
    ws.merge_cells(f"H{start_row}:H{start_row+4}")

    # amount after discount（列 9）：合并 5 行
    ws.merge_cells(f"I{start_row}:I{start_row+4}")

    # delivery（列 10）：合并前 2 行（行 1-2）
    ws.merge_cells(f"J{start_row}:J{start_row+1}")


def extract_existing_data(ws) -> List[Tuple[int, Dict[str, Any]]]:
    """
    从现有工作表中提取已有数据

    Args:
        ws: 工作表对象

    Returns:
        列表，每个元素是 (序号, 数据字典) 的元组
    """
    existing_data = []
    row_idx = 2  # 从第 2 行开始（第 1 行是表头）

    while row_idx <= ws.max_row:
        # 检查是否有数据（序号列不为空）
        seq_cell = ws.cell(row=row_idx, column=1)
        if seq_cell.value is None or str(seq_cell.value).strip() == "":
            break

        # 读取第 1 行数据（主数据）
        item = {
            "序号": seq_cell.value,
            "PO": ws.cell(row=row_idx, column=2).value or "",
            "style no": ws.cell(row=row_idx, column=3).value or "",
            "style name": ws.cell(row=row_idx, column=4).value or "",
            "color": ws.cell(row=row_idx, column=5).value or "",
            "unit price": ws.cell(row=row_idx, column=6).value or "",
            "quantity": ws.cell(row=row_idx, column=7).value or "",
            "total": ws.cell(row=row_idx, column=8).value or "",
            "Ex-date": ws.cell(row=row_idx, column=10).value or "",
            "原PDF名称": ws.cell(row=row_idx, column=11).value or "",
        }

        # 读取第 2 行（style code）
        item["style code"] = ws.cell(row=row_idx + 1, column=2).value or ""

        # 读取第 3 行（color code）
        item["color code"] = ws.cell(row=row_idx + 2, column=2).value or ""

        # 读取第 5 行（sell）
        item["sell"] = ws.cell(row=row_idx + 4, column=2).value or ""

        existing_data.append((int(item["序号"]), item))
        row_idx += 6  # 跳过 6 行（一组数据）

    return existing_data


def fill_data_to_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, data: List[Dict[str, Any]], start_idx: int = 0) -> None:
    """
    将数据填充到工作表，并合并单元格

    Args:
        ws: 工作表对象
        data: 要填充的数据列表
        start_idx: 起始序号
    """
    for idx, item in enumerate(data):
        # 计算起始行（每组数据占 6 行，从第 2 行开始）
        start_row = 2 + ((start_idx + idx) * 6)
        seq_num = start_idx + idx + 1  # 序号

        # 获取数据值，如果不存在则为空字符串
        po = str(item.get("PO", ""))
        style_no = str(item.get("style no", ""))
        style_name = str(item.get("style name", ""))
        color = str(item.get("color", ""))
        unit_price = str(item.get("unit price", ""))
        quantity = str(item.get("quantity", ""))
        total = str(item.get("total", ""))
        ex_date = str(item.get("Ex-date", ""))
        style_code = str(item.get("style code", ""))
        color_code = str(item.get("color code", ""))
        sell = str(item.get("sell", ""))
        pdf_name = str(item.get("原PDF名称", ""))

        # 添加前缀
        if po:
            po = f"PO#{po}"
        if style_code:
            style_code = f"style code#{style_code}"
        if color_code:
            color_code = f"color code#{color_code}"
        if sell:
            sell = f"GBP{sell}"
        if total:
            total = f"¥{total}"

        # 填充第 1 行（序号 + 主数据）
        ws.cell(row=start_row, column=1, value=seq_num)
        ws.cell(row=start_row, column=2, value=po)
        ws.cell(row=start_row, column=3, value=style_no)
        ws.cell(row=start_row, column=4, value=style_name)
        ws.cell(row=start_row, column=5, value=color)
        ws.cell(row=start_row, column=6, value=unit_price)
        ws.cell(row=start_row, column=7, value=quantity)
        ws.cell(row=start_row, column=8, value=total)
        ws.cell(row=start_row, column=9, value="")
        ws.cell(row=start_row, column=10, value=ex_date)
        ws.cell(row=start_row, column=11, value=pdf_name)

        # 填充第 2 行（style code）
        ws.cell(row=start_row + 1, column=2, value=style_code)

        # 填充第 3 行（color code）
        ws.cell(row=start_row + 2, column=2, value=color_code)

        # 填充第 4 行（空行）
        ws.cell(row=start_row + 3, column=2, value="")
        ws.cell(row=start_row + 3, column=5, value="")
        ws.cell(row=start_row + 3, column=6, value="")
        ws.cell(row=start_row + 3, column=8, value="")
        ws.cell(row=start_row + 3, column=9, value="")
        ws.cell(row=start_row + 3, column=10, value="")
        ws.cell(row=start_row + 3, column=11, value="")

        # 填充第 5 行（sell）
        ws.cell(row=start_row + 4, column=2, value=sell)

        # 第 6 行保持空（分隔行）

        # 合并单元格
        merge_cells_for_group(ws, start_row)


def fill_template(data: List[Dict[str, Any]], template_path: str, output_path: str) -> None:
    """
    基于模板填充数据（新建模式）

    Args:
        data: 要填充的数据列表
        template_path: 模板文件路径
        output_path: 输出 Excel 文件路径
    """
    if not data:
        raise ValueError("没有数据可填充")

    # 复制模板文件
    shutil.copy(template_path, output_path)

    # 打开工作簿
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # 清空模板中的示例数据（如果有）
    if ws.max_row > 1:
        # 删除所有数据行
        for row_idx in range(ws.max_row, 1, -1):
            ws.delete_rows(row_idx)

    # 填充数据
    fill_data_to_sheet(ws, data)

    # 保存 Excel 文件
    wb.save(output_path)
    print(f"Excel 文件已成功导出到: {output_path}")
    print(f"共填充 {len(data)} 组数据（每组 6 行）")
    print(f"总行数: {2 + len(data) * 6}")


def update_excel(new_data: List[Dict[str, Any]], template_path: str, excel_path: str) -> None:
    """
    更新 Excel 文件（增量更新模式）

    Args:
        new_data: 新的数据列表
        template_path: 模板文件路径
        excel_path: Excel 文件路径
    """
    if not new_data:
        raise ValueError("没有数据可更新")

    # 检查 Excel 文件是否存在
    if not os.path.exists(excel_path):
        # 不存在，基于模板创建
        print(f"Excel 文件不存在，基于模板创建: {excel_path}")
        fill_template(new_data, template_path, excel_path)
        return

    print(f"Excel 文件已存在，执行增量更新: {excel_path}")

    # 读取现有 Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 提取现有数据
    existing_data = extract_existing_data(ws)
    print(f"现有数据: {len(existing_data)} 组")

    # 获取新数据的 PDF 文件名集合
    new_pdf_names = {item.get("原PDF名称", "") for item in new_data}

    # 过滤掉与新数据 PDF 文件名相同的旧数据
    filtered_data = []
    for seq_num, item in existing_data:
        if item.get("原PDF名称", "") not in new_pdf_names:
            filtered_data.append(item)

    removed_count = len(existing_data) - len(filtered_data)
    if removed_count > 0:
        print(f"删除了 {removed_count} 组重复数据")

    # 合并数据：先保留旧数据，再追加新数据
    merged_data = filtered_data + new_data

    # 清空工作表数据（保留表头）
    for row_idx in range(ws.max_row, 1, -1):
        ws.delete_rows(row_idx)

    # 重新填充所有数据
    fill_data_to_sheet(ws, merged_data)

    # 保存 Excel 文件
    wb.save(excel_path)
    print(f"Excel 文件已成功更新: {excel_path}")
    print(f"当前共有 {len(merged_data)} 组数据")


def load_json_data(json_path: str) -> List[Dict[str, Any]]:
    """
    从 JSON 文件加载数据

    Args:
        json_path: JSON 文件路径

    Returns:
        数据列表
    """
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"JSON 文件不存在: {json_path}")

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 确保数据是列表格式
    if isinstance(data, dict):
        # 如果是单个字典，包装成列表
        data = [data]
    elif not isinstance(data, list):
        raise ValueError("JSON 数据格式不正确，应该是对象或对象数组")

    return data


def main():
    """主函数"""
    if len(sys.argv) < 4:
        print("使用方法: python export_to_excel.py <json_data_path> <template_path> <excel_output_path>")
        print("")
        print("参数说明:")
        print("  json_data_path      - 包含提取字段数据的 JSON 文件路径")
        print("  template_path       - Excel 模板文件路径")
        print("  excel_output_path   - 输出的 Excel 文件路径（如果已存在则自动更新）")
        print("")
        print("示例:")
        print("  python export_to_excel.py ./extracted_data.json ./template.xlsx ./output.xlsx")
        print("")
        print("字段映射关系:")
        for old_key, new_key in FIELD_MAPPING.items():
            print(f"  {old_key} → {new_key}")
        print("")
        print("字段值前缀规则（导出时自动添加）:")
        print("  PO → PO#PO值")
        print("  style code → style code#style code值")
        print("  color code → color code#color code值")
        print("  sell → GBPsell值")
        print("  total → ¥total值")
        sys.exit(1)

    json_path = sys.argv[1]
    template_path = sys.argv[2]
    excel_path = sys.argv[3]

    try:
        # 加载数据
        print(f"正在加载数据: {json_path}")
        data = load_json_data(json_path)
        print(f"成功加载 {len(data)} 条记录")

        # 映射字段（如果数据中包含"原PDF名称"字段，会保留）
        print("正在映射字段...")
        mapped_data = map_fields(data)

        # 更新 Excel（自动判断是新建还是更新）
        update_excel(mapped_data, template_path, excel_path)

        return 0

    except FileNotFoundError as e:
        print(f"错误: {e}")
        return 1
    except json.JSONDecodeError as e:
        print(f"JSON 解析错误: {e}")
        return 1
    except ValueError as e:
        print(f"数据错误: {e}")
        return 1
    except Exception as e:
        print(f"导出过程中发生错误: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
